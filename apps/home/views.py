from django import template
from django.http import Http404, HttpResponseForbidden, JsonResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from apps.home.models import *
from apps.home.utils.reports import generate_farmer_report_pdf, generate_vet_report_excel
from .forms import *
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from apps.home.models import *
from django.db.models import Count
from django_daraja.mpesa.core import MpesaClient
from django.utils.timezone import now
from decimal import Decimal, InvalidOperation
from .mpesa_utils import send_stk_push
import json
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side




# @login_required(login_url="/login/")
# def index(request):
#     context = {'segment': 'index'}

#     html_template = loader.get_template('home/index.html')
#     return HttpResponse(html_template.render(context, request))


@login_required(login_url="/login/")
def index(request):
    user_role = getattr(request.user, "role", None)

    if user_role == 'admin':
        return redirect('admin_dashboard')
    
    elif user_role == 'staff':
        return redirect('staff_dashboard')
    
    elif user_role == 'farmer':
        return redirect('farmer_dashboard')
    

    elif user_role == 'vet':
        return redirect('vet_dashboard')
    

    else:
        return render(request, "home/page-403.html") 

@login_required(login_url="/login/")

def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))



# Admin Dashboard View
@login_required(login_url="/login/")
def admin_dashboard(request):
    # return render(request, "admin/admin_dashboard.html")
    return render(request, "home/index.html")


class AppointmentListView(LoginRequiredMixin, ListView):
    model = Appointment
    template_name = "vet/appointments.html"
    context_object_name = "appointments"
    ordering = ['-date']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Check if user is a vet or has staff permissions
        if self.request.user.is_staff or getattr(self.request.user, "role", "") == "vet":
            return queryset  # Show all appointments
        
        # Otherwise, filter for farmer appointments only
        return queryset.filter(farmer=self.request.user)


# class AppointmentCreateView(LoginRequiredMixin, CreateView):
#     model = Appointment
#     form_class = AppointmentForm
#     template_name = "vet/appointment_form.html"
#     success_url = reverse_lazy("appointments")
class AppointmentCreateView(LoginRequiredMixin, CreateView):
    model = Appointment
    form_class = AppointmentForm
    template_name = "vet/appointment_form.html"
    success_url = reverse_lazy("appointments")

    def get_form_kwargs(self):
        """Pass the current user to the form so we can filter livestock."""
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)

        # If the logged-in user is a farmer, hide farmer field & prefill it
        if getattr(self.request.user, "role", "") == "farmer":
            form.fields["farmer"].widget = forms.HiddenInput()
            form.fields["farmer"].initial = self.request.user

        return form

    def form_valid(self, form):
        """Ensure farmer is set automatically for farmers."""
        if getattr(self.request.user, "role", "") == "farmer":
            form.instance.farmer = self.request.user
        return super().form_valid(form)

    def form_valid(self, form):
        # For farmers, force the logged-in user as the farmer
        if getattr(self.request.user, "role", "") == "farmer":
            form.instance.farmer = self.request.user
        return super().form_valid(form)

class AppointmentUpdateView(LoginRequiredMixin, UpdateView):
    model = Appointment
    form_class = AppointmentForm
    template_name = "vet/appointment_form.html"
    success_url = reverse_lazy("appointments")

    def get_object(self, queryset=None):
        """Restrict farmers to editing only their own appointments."""
        obj = super().get_object(queryset)
        if getattr(self.request.user, "role", "") == "farmer" and obj.farmer != self.request.user:
            raise Http404("You are not allowed to edit this appointment.")
        return obj

    def get_form_kwargs(self):
        """Pass current user to the form to filter livestock."""
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)

        # If farmer is editing, hide farmer field and lock it to their account
        if getattr(self.request.user, "role", "") == "farmer":
            form.fields["farmer"].widget = forms.HiddenInput()
            form.fields["farmer"].initial = self.request.user

        return form

class AppointmentDeleteView(LoginRequiredMixin, DeleteView):
    model = Appointment
    template_name = "vet/appointment_confirm_delete.html"
    success_url = reverse_lazy("appointments")

    def get_object(self, queryset=None):
        """Restrict farmers from deleting other farmers' appointments."""
        obj = super().get_object(queryset)
        if getattr(self.request.user, "role", "") == "farmer" and obj.farmer != self.request.user:
            raise Http404("You are not allowed to delete this appointment.")
        return obj


# Staff Dashboard View
@login_required(login_url="/login/")
def staff_dashboard(request):
    # livestock = Livestock.objects.all()
    # return render(request, "staff/staff_dashboard.html",{'livestock': livestock})
    user = request.user

    if user.is_staff:  # Admin can see all livestock
        total_livestock = Livestock.objects.count()
    elif hasattr(user, 'is_vet') and user.is_vet:  # Vet sees livestock assigned to them
        total_livestock = Livestock.objects.filter(assigned_vet=user).count()
    else:  # Farmer sees only their livestock
        total_livestock = Livestock.objects.filter(owner=user).count()

    return render(request,  "staff/staff_dashboard.html", {'total_livestock': total_livestock})



@login_required(login_url="/login/")
def farmer_dashboard(request):
    user = request.user

    # Determine which livestock to show based on role
    if user.role == 'admin':
        livestock_qs = Livestock.objects.all()
    elif user.role == 'vet':
        livestock_qs = Livestock.objects.all()
    elif user.role == 'farmer':
        livestock_qs = Livestock.objects.filter(farmer=user)
    else:
        return render(request, "home/page-403.html")

    # Count livestock types
    # livestock_counts = livestock_qs.values("livestock_type").annotate(total=Count("id"))
    cattle_list = livestock_qs.filter(livestock_type='cattle')
    livestock_counts = (livestock_qs.values("livestock_type").annotate(total=Count("id")))
    
    recent_vaccinations = Vaccination.objects.filter(
            livestock__farmer=user
        ).order_by("-vaccination_date")[:5]
    # Prepare health status per animal
    livestock_list = []
    for animal in livestock_qs:
        due_vaccinations = animal.vaccination_set.filter(next_due_date__lte=now()).exists()
        livestock_list.append({
            "name": animal.name,
            "type": animal.livestock_type,
            "weight": animal.weight,
            "breed": animal.breed,
            "health_status": "Vaccination due" if due_vaccinations else "Healthy",
        })
        
    recent_treatments = Treatment.objects.filter(
        livestock__farmer=user
    ).order_by("-treatment_date")[:5]
    context = {
        # "total_livestock": livestock_qs.count(),
        
        "cattle_list": cattle_list,
        "total_livestock": livestock_qs.count(),
        "total_cattle": next((x["total"] for x in livestock_counts if x["livestock_type"] == "cattle"), 0),
        "total_sheep": next((x["total"] for x in livestock_counts if x["livestock_type"] == "sheep"), 0),
        "total_goats": next((x["total"] for x in livestock_counts if x["livestock_type"] == "goat"), 0),
        "total_poultry": next((x["total"] for x in livestock_counts if x["livestock_type"] == "poultry"), 0),
        "total_other": next((x["total"] for x in livestock_counts if x["livestock_type"] == "other"), 0),
        "livestock_list": livestock_list,
        "recent_vaccinations": recent_vaccinations,
        "recent_treatments":recent_treatments,
    }

    return render(request, "farmer/farmer_dashboard.html", context)


class FarmerTreatmentListView(LoginRequiredMixin, ListView):
    model = Treatment
    template_name = 'farmer/farmer_treatment_list.html'  # Create this template
    context_object_name = 'treatments'

    def get_queryset(self):
        # Only show treatments for livestock that belongs to the logged-in farmer
        return Treatment.objects.filter(livestock__farmer=self.request.user).order_by('-treatment_date')




@login_required
def livestock_list(request):
    user = request.user
    query = request.GET.get('q')  # Get search query from the URL

    # Determine which livestock to show based on role
    if user.role in ['admin', 'vet']:
        livestock = Livestock.objects.all()
    elif user.role == 'farmer':
        livestock = Livestock.objects.filter(farmer=user)
    else:
        livestock = Livestock.objects.none()  # no access for other roles

    # Apply search filtering if query is provided
    if query:
        livestock = livestock.filter(
            Q(name__icontains=query) |
            Q(livestock_type__icontains=query) |   # ✅ replaced tag with livestock_type
            Q(breed__icontains=query)
        )

    # Handle form submission
    if request.method == "POST":
        form = LivestockForm(request.POST)
        if form.is_valid():
            livestock_instance = form.save(commit=False)
            # Only set farmer automatically if user is a farmer
            if user.role == 'farmer':
                livestock_instance.farmer = user
            livestock_instance.save()
            messages.success(request, "Livestock added successfully!")
            return redirect('livestock_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LivestockForm()

    return render(request, 'farmer/livestock.html', {
        'livestock': livestock,
        'form': form,
        'query': query
    })


@login_required
def vet_livestock_list(request):
    user = request.user

    # Determine which livestock to show based on role
    if user.role in ['admin', 'vet']:
        livestock = Livestock.objects.all()
    elif user.role == 'farmer':
        livestock = Livestock.objects.filter(farmer=user)
    else:
        livestock = Livestock.objects.none()  # no access for other roles

    if request.method == "POST":
        form = LivestockForm(request.POST)
        if form.is_valid():
            livestock_instance = form.save(commit=False)
            
            # Only set farmer automatically if user is a farmer
            if user.role == 'farmer':
                livestock_instance.farmer = user
            
            livestock_instance.save()
            messages.success(request, "Livestock added successfully!")
            return redirect('livestock_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LivestockForm()

    return render(request, 'vet/livestock.html', {
        'livestock': livestock,
        'form': form,
    })

@login_required
def livestock_edit(request, pk):
    livestock_instance = get_object_or_404(Livestock, pk=pk, farmer=request.user)

    if request.method == "POST":
        form = LivestockForm(request.POST, instance=livestock_instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Livestock updated successfully!")
            return redirect('livestock_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = LivestockForm(instance=livestock_instance)

    return render(request, 'farmer/livestock_edit.html', {'form': form})


@login_required
def livestock_delete(request, pk):
    livestock_instance = get_object_or_404(Livestock, pk=pk, farmer=request.user)

    if request.method == "POST":
        livestock_instance.delete()
        messages.success(request, "Livestock deleted successfully!")
        return redirect('livestock_list')

    return render(request, 'farmer/livestock_delete.html', {'livestock': livestock_instance})




def vaccination_list(request):
    user = request.user

    # Admins see all vaccinations
    if user.is_superuser:
        vaccinations = Vaccination.objects.all()

    # Farmers see only vaccinations of their livestock
    elif hasattr(user, 'farmer'):
        vaccinations = Vaccination.objects.filter(livestock__farmer=user.farmer)

    # Vets see only vaccinations of livestock they're assigned to
    elif hasattr(user, 'vet'):
        vaccinations = Vaccination.objects.filter(livestock__appointment__vet=user.vet).distinct()

    else:
        vaccinations = Vaccination.objects.none()

    return render(request, "farmer/vaccination_list.html", {"vaccinations": vaccinations})

@login_required(login_url="/login/")
def vet_dashboard(request):
    user = request.user

    if user.role != "vet":
        return render(request, "home/page-403.html")

    # Show all livestock
    livestock_qs = Livestock.objects.all()

    # Count livestock types
    livestock_counts = livestock_qs.values("livestock_type").annotate(total=Count("id"))

    # Total stats
    total_vaccinations = Vaccination.objects.count()
    total_treatments = Treatment.objects.count()
    total_appointments = Appointment.objects.count()

    # Prepare health status for all livestock
    livestock_list = []
    for animal in livestock_qs:
        due_vaccinations = animal.vaccination_set.filter(next_due_date__lte=now()).exists()
        livestock_list.append({
            "name": animal.name,
            "type": animal.livestock_type,
            "weight": animal.weight,
            "breed": animal.breed,
            "health_status": "Vaccination due" if due_vaccinations else "Healthy",
        })

    recent_vaccinations = Vaccination.objects.all().order_by("-vaccination_date")[:5]
    recent_treatments = Treatment.objects.all().order_by("-treatment_date")[:5]
    recent_appointments = Appointment.objects.all().order_by("-date")[:5]

    context = {
        "total_livestock": livestock_qs.count(),
        "total_vaccinations": total_vaccinations,
        "total_treatments": total_treatments,
        "total_appointments": total_appointments,
        "livestock_list": livestock_list,
        "recent_vaccinations": recent_vaccinations,
        "recent_treatments": recent_treatments,
        "recent_appointments": recent_appointments,
        "total_cattle": next((x["total"] for x in livestock_counts if x["livestock_type"] == "cattle"), 0),
        "total_sheep": next((x["total"] for x in livestock_counts if x["livestock_type"] == "sheep"), 0),
        "total_goats": next((x["total"] for x in livestock_counts if x["livestock_type"] == "goat"), 0),
        "total_poultry": next((x["total"] for x in livestock_counts if x["livestock_type"] == "poultry"), 0),
        "total_other": next((x["total"] for x in livestock_counts if x["livestock_type"] == "other"), 0),
    }

    return render(request, "vet/vet_dashboard.html", context)


class VetTreatmentListView(LoginRequiredMixin, ListView):
    model = Treatment
    template_name = "vet/treatment_list.html"   # your template file
    context_object_name = "treatments"
    paginate_by = 10  # show 10 per page

    def get_queryset(self):
        return (
            Treatment.objects
            .filter(vet=self.request.user)
            .select_related('livestock')        # if you want livestock fields
            .order_by('-treatment_date')
       
)


class VetTreatmentUpdateView(LoginRequiredMixin, UpdateView):
    model = Treatment
    form_class = TreatmentForm
    template_name = "vet/treatment_form.html"
    success_url = reverse_lazy("vet_treatments")

class VetTreatmentDeleteView(LoginRequiredMixin, DeleteView):
    model = Treatment
    template_name = "vet/treatment_confirm_delete.html"
    success_url = reverse_lazy("vet_treatments")    


@login_required
def vaccination_list(request):
    user = request.user

    # If admin or staff → show all vaccinations
    if user.is_staff or user.role == "admin" or user.role == "staff":
        vaccinations = Vaccination.objects.all()

    # If vet → show vaccinations assigned to them
    elif user.role == "vet":
        vaccinations = Vaccination.objects.filter(vet=user)

    # Otherwise, assume farmer → show vaccinations for farmer's livestock only
    else:
        vaccinations = Vaccination.objects.filter(livestock__farmer=user)

    # Show "Add Vaccination" button only for vet, admin, and staff
    show_add_button = (
        user.is_staff
        or user.role == "vet"
        or user.role == "admin"
        or user.role == "staff"
    )

    return render(
        request,
        "vet/vaccination_list.html",
        {
            "vaccinations": vaccinations,
            "show_add_button": show_add_button,
        },
    )

class VaccinationCreateView(LoginRequiredMixin, CreateView):
    model = Vaccination
    form_class = VaccinationForm
    template_name = "vet/vaccination_form.html"
    success_url = reverse_lazy("vaccination_list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user  # Pass logged-in user to form
        return kwargs

    def form_valid(self, form):
        # Assign the logged-in vet automatically
        form.instance.vet = self.request.user
        return super().form_valid(form)
   

class VaccinationUpdateView(LoginRequiredMixin, UpdateView):
    model = Vaccination
    form_class = VaccinationForm
    template_name = "vet/vaccination_form.html"
    success_url = reverse_lazy("vaccination_list")

class VaccinationDeleteView(LoginRequiredMixin, DeleteView):
    model = Vaccination
    template_name = "vet/vaccination_confirm_delete.html"
    success_url = reverse_lazy("vaccination_list")




@login_required
def farmer_report(request):
    user = request.user

    # Show only farmer's livestock if farmer; otherwise, show all.
    if user.role == "farmer":
        livestock = Livestock.objects.filter(farmer=user)
    else:
        livestock = Livestock.objects.all()

    # Start with all vaccinations related to livestock
    vaccinations = Vaccination.objects.filter(livestock__in=livestock)

    # Get filters from request
    livestock_id = request.GET.get("livestock")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    # Filter by livestock if selected
    if livestock_id:
        vaccinations = vaccinations.filter(livestock_id=livestock_id)

    # Filter by date range if provided
    if from_date:
        vaccinations = vaccinations.filter(vaccination_date__gte=parse_date(from_date))
    if to_date:
        vaccinations = vaccinations.filter(vaccination_date__lte=parse_date(to_date))

    # If "download" is requested → Generate PDF
    if request.GET.get("download") == "pdf":
        return generate_vaccination_report_pdf(vaccinations, livestock_id, from_date, to_date)

    # Pass context to template
    context = {
        "livestock": livestock,
        "vaccinations": vaccinations,
    }
    return render(request, "farmer/report.html", context)


def generate_vaccination_report_pdf(vaccinations, livestock_id=None, from_date=None, to_date=None):
    """Generate vaccination report as PDF"""
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="vaccination_report.pdf"'

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("🐄 Farmer Vaccination Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Filters summary
    filters_summary = []
    if livestock_id:
        livestock = Livestock.objects.get(pk=livestock_id)
        filters_summary.append(f"<b>Livestock:</b> {livestock.name} ({livestock.livestock_type})")
    if from_date and to_date:
        filters_summary.append(f"<b>Date Range:</b> {from_date} → {to_date}")
    elif from_date:
        filters_summary.append(f"<b>From:</b> {from_date}")
    elif to_date:
        filters_summary.append(f"<b>Up to:</b> {to_date}")

    if filters_summary:
        elements.append(Paragraph(", ".join(filters_summary), styles["Normal"]))
        elements.append(Spacer(1, 12))

    # Table header
    data = [["Livestock", "Vaccine", "Date"]]
    for v in vaccinations:
        data.append([
            v.livestock.name,
            v.vaccine_name,
            v.vaccination_date.strftime("%d-%m-%Y"),
        ])

    # If no data, show message
    if len(data) == 1:
        data.append(["No records found", "", ""])

    # Table styling
    table = Table(data, colWidths=[150, 150, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(table)

    # Build PDF
    doc.build(elements)
    return response

from openpyxl.utils import get_column_letter
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from datetime import datetime
@login_required
def vet_report(request):
    user = request.user

    # Fetch appointments for this vet only
    appointments = Appointment.objects.filter(vet=user).order_by('-date')

    # Get filters from request safely
    from_date_raw = request.GET.get("from_date")
    to_date_raw = request.GET.get("to_date")

    from_date = parse_date(from_date_raw) if from_date_raw and from_date_raw != "None" else None
    to_date = parse_date(to_date_raw) if to_date_raw and to_date_raw != "None" else None

    # Convert to timezone-aware datetimes
    if from_date:
        from_date = make_aware(datetime.combine(from_date, datetime.min.time()))
        appointments = appointments.filter(date__gte=from_date)
    if to_date:
        to_date = make_aware(datetime.combine(to_date, datetime.max.time()))
        appointments = appointments.filter(date__lte=to_date)

    # If "download" is requested → Generate Excel
    if request.GET.get("download") == "excel":
        return generate_appointment_report_excel(appointments, from_date_raw, to_date_raw)

    # Pass context to template
    context = {
        "appointments": appointments,
        "from_date": from_date_raw if from_date_raw != "None" else "",
        "to_date": to_date_raw if to_date_raw != "None" else "",
    }
    return render(request, "vet/report.html", context)



def generate_appointment_report_excel(appointments, from_date=None, to_date=None):
    """Generate vet appointments report as an Excel file"""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="vet_appointments_report.xlsx"'

    # Create workbook & worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Vet Appointments Report"

    # Title row
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "📅 Veterinary Appointments Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # Filter summary row
    ws.merge_cells("A2:D2")
    filter_cell = ws["A2"]
    if from_date and to_date:
        filter_cell.value = f"Date Range: {from_date} → {to_date}"
    elif from_date:
        filter_cell.value = f"From: {from_date}"
    elif to_date:
        filter_cell.value = f"Up to: {to_date}"
    else:
        filter_cell.value = "All Appointments"
    filter_cell.font = Font(italic=True, size=12)
    filter_cell.alignment = Alignment(horizontal="center")

    # Column headers
    headers = ["Date", "Farmer", "Livestock", "Purpose"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for appt in appointments:
        ws.append([
            appt.date.strftime("%d-%m-%Y %H:%M"),
            appt.farmer.username,
            appt.livestock.name,
            appt.purpose,
        ])

    # If no data, add empty row
    if not appointments.exists():
        ws.append(["No appointments found", "", "", ""])

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    wb.save(response)
    return response


# from .mpesa import send_stk_push  # your helper


NGROK_BASE = "https://b08250fcf4c6.ngrok-free.app"  
# MPESA callback received: {'Body': {'stkCallback': 
#                                    {'MerchantRequestID': '49ff-4efd-ad2f-9dc3e15e6f9b14671', 
#                                     'CheckoutRequestID': 'ws_CO_17092025141218718727824180', 
#                                     'ResultCode': 1032, 
#                                     'ResultDesc': 'Request Cancelled by user.'}}
#                                     }
def add_treatment(request):
    if request.method == "POST":
        print("=== add_treatment view called === POST")
        form = TreatmentForm(request.POST)
        if form.is_valid():
            # Save the Treatment but don’t commit yet
            treatment = form.save(commit=False)
            treatment.vet = request.user  # link vet to treatment
            treatment.save()
            print(f"Saved treatment id: {treatment.id}")

            # prepare phone for STK push (but not saved in Treatment)
            raw_phone = form.cleaned_data.get("phone_number")  # must exist in your form
            cleaned_phone = None
            if raw_phone:
                cleaned_phone = raw_phone.strip().replace("+", "").replace(" ", "")
                if cleaned_phone.startswith("0"):
                    cleaned_phone = "254" + cleaned_phone[1:]

            # send STK push only if we have phone and cost
            try:
                raw_cost = form.cleaned_data.get("cost")
                if raw_cost is None:
                    raise ValueError("Cost missing in form")
                if not cleaned_phone:
                    raise ValueError("Phone number missing in form")

                amount = int(Decimal(raw_cost))  # must be int for M-Pesa
                account_ref = f"Treatment{treatment.id}"
                desc = "Vet treatment payment"
                callback_url = NGROK_BASE + reverse("mpesa_callback")  # 👈 public URL

                print("About to send STK Push…")
                response = send_stk_push(
                    phone=cleaned_phone,
                    amount=amount,
                    account_reference=account_ref,
                    transaction_desc=desc,
                    callback_url=callback_url,
                )
                print("STK push status:", response.status_code)
                print("STK push body:", response.text)

                # if response.status_code == 200:
                #     messages.success(
                #         request,
                #         "Treatment saved and STK push initiated. Enter PIN on your phone."
                #     )
                # else:
                #     messages.warning(
                #         request,
                #         f"Treatment saved but STK push failed: {response.text}"
                #     )
                if response.status_code == 200:
                    data = response.json()
                    # 🔽 store checkout id on the treatment record
                    treatment.checkout_request_id = data.get("CheckoutRequestID")
                    treatment.save(update_fields=["checkout_request_id"])
                    messages.success(
                        request,
                        "Treatment saved and STK push initiated. Enter PIN on your phone."
                    )

                else:
                    messages.warning(
                        request,
                        f"Treatment saved but STK push failed: {response.text}"
                    )

            except (InvalidOperation, Exception) as e:
                print("STK push failed:", e)
                messages.warning(
                    request,
                    f"Treatment saved but STK push failed: {e}"
                )

            return redirect("vet_treatments")  # your list URL name
        else:
            print("Form errors:", form.errors)
    else:
        form = TreatmentForm()

    return render(request, "vet/treatment_form.html", {"form": form})

    
from django.views.decorators.csrf import csrf_exempt





@csrf_exempt
def mpesa_callback(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
        print("MPESA callback received:", data)

        stk = data.get("Body", {}).get("stkCallback", {})
        result_code = stk.get("ResultCode")
        result_desc = stk.get("ResultDesc")
        checkout_id = stk.get("CheckoutRequestID")

        # Find treatment by checkout_request_id
        treatment = Treatment.objects.filter(checkout_request_id=checkout_id).first()
        if treatment:
            # Save full callback
            treatment.mpesa_callback_body = data
            treatment.mpesa_result_desc = result_desc

            if result_code == 0:  # Payment success
                items = stk.get("CallbackMetadata", {}).get("Item", [])
                meta = {i["Name"]: i.get("Value") for i in items}
                treatment.mpesa_receipt_number = meta.get("MpesaReceiptNumber")
                treatment.mpesa_phone_number = meta.get("PhoneNumber")
                treatment.mpesa_transaction_date = str(meta.get("TransactionDate"))
                treatment.status = "paid"

            treatment.save()
            print(f"Treatment {treatment.id} updated with MPESA payment info.")
        else:
            print(f"No treatment found for CheckoutRequestID {checkout_id}")

    except Exception as e:
        print("Error processing MPESA callback:", e)

    return JsonResponse({"ResultCode": 0, "ResultDesc": "Accepted"})
